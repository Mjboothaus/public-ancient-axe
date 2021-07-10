import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# Streamlit Page Config has to be first thing in script

st.set_page_config(
    page_title="old-ancient-axe",
    page_icon=None,
    layout="centered",
    initial_sidebar_state="auto",
)

# Local functions

def get_sheet_names(wb):
    if wb is None:
        return ['No Excel workbook is loaded']
    else:
        return [sheetname for sheetname in wb.sheetnames]

def extract_named_tables(sheetname, wb):
    if wb is None:
        return []
    sheet = wb[sheetname]
    #if sheet.tables.keys() is not None:
    return sheet.tables.items()
    #else:
    #    return "No named tables in sheet"

def get_names_in_sheet(sheetname, wb):
    if wb is None:
        return []
    sheetid = wb.sheetnames.index(sheetname)
    tmp = [[defined_name.name, defined_name.value] for defined_name in wb.defined_names.localnames(sheetid)]
    return tmp

    #return wb.defined_names.localnames(sheetid)

# Define sidebar class (for ease of tracking/grouping controls)
class SideBar:
    uploaded_file = st.empty()
    workbook_name = ""
    sheet_select = st.empty()


# Start App layout

st.title("No Data in Excel - Prototype App")

sb = SideBar()
sheetnames = ['Please upload an Excel workbook']

sb.uploaded_file = st.sidebar.file_uploader("Upload Files", type=["xlsx"])

wb = None

if sb.uploaded_file is not None:
    try:
        wb = load_workbook(filename=sb.uploaded_file)
        st.write(sb.uploaded_file.name)
    except:
        st.info("No Excel file uploaded")
   

#    file_details = {
#        "FileName": uploaded_file.name,
#        "FileType": uploaded_file.type,
#        "FileSize": uploaded_file.size,
#    }
#    st.write(file_details)

sheetnames = get_sheet_names(wb)

sb.sheet_select = st.sidebar.radio('Worksheets:', sheetnames, 0)

st.header(sb.sheet_select)
st.write(get_names_in_sheet(sb.sheet_select, wb))
st.write(extract_named_tables(sb.sheet_select, wb))



# st.write(named_tables)

# See https://databooth.slite.com/api/s/note/KFoQMJupyw4ix32aSHqvdr/old-ancient-axe-No-Data-in-Excel-app-ideas